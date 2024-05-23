from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QLabel, QLineEdit, QPushButton, QVBoxLayout, QWidget

from ErrorWindow import ErrorWindow
from Resources import RESOURCES
from Functions import GetExcelFileName, AddingFileExt, IfValue, DATETIME
from AfterMerge import AfterMerge

import os
import pathlib

class MergeWindow(ErrorWindow, RESOURCES):
    def __init__(self):
        super().__init__()
        
        self.setWindowTitle("Merge Window")
        self.setFixedSize(250, 100)
        
        self.error = ErrorWindow()
        self.AfterMergeWindow = AfterMerge()
        
        RESOURCES.buildFileNameNoExt = "Final_Bank_Statement"
        RESOURCES.buildFileNameExt = AddingFileExt(self.buildFileNameNoExt)

        layout = QVBoxLayout()

        self.buildFileLabel = QLabel("Build File Name - {}(Default)".format(RESOURCES.buildFileNameNoExt))
        self.buildFileInput = QLineEdit()
        self.mergeButton = QPushButton("Merge")
        self.container = QWidget()
        
        self.buildFileInput.setPlaceholderText("Enter Build File Name")

        layout.addWidget(self.buildFileLabel)
        layout.addWidget(self.buildFileInput)
        layout.addWidget(self.mergeButton)
        
        self.container.setLayout(layout)
        self.setCentralWidget(self.container)
        
        self.buildFileInput.textChanged.connect(self.ChangedBuildFileName)
        self.mergeButton.clicked.connect(self.FinalMerge)
    
    def ChangedBuildFileName(self):
        RESOURCES.buildFileNameNoExt = self.buildFileInput.text()
        RESOURCES.buildFileNameExt = AddingFileExt(RESOURCES.buildFileNameNoExt)
        self.buildFileLabel.setText("Build File Name - {}".format(RESOURCES.buildFileNameNoExt))
        

    def FinalMerge(self):
        numExcelToMerge = len(RESOURCES.excelList)
        rawBuildPath = pathlib.PureWindowsPath(os.getcwd())
        pureBuildPath = rawBuildPath.as_posix()
        
        if(numExcelToMerge == 0):
            self.error.show()
            ErrorWindow.label.setText("Files to be Merged are not Provided")
        
        elif(self.buildFileNameNoExt.isspace() == True or RESOURCES.buildFileNameNoExt == ""):
            self.error.show()
            ErrorWindow.label.setText("Build File Needs to have a Name")

        elif(os.path.isfile(pureBuildPath + "/" + RESOURCES.buildFileNameExt)):
        
            self.error.show()
            ErrorWindow.label.setText("Another File with the Same Name Already Exists")

        else:
            
            #Creating build excel
            docBuildWorkbook = Workbook()
            docBuildSheet = docBuildWorkbook.active
            docBuildWorkbook.save(filename=RESOURCES.buildFileNameExt)
        
            #Merge loop
            DATA = []
            numExcelMerged = 0
            headingValue = 0
            
            while(numExcelMerged != numExcelToMerge):
    
                #Opening an excel
                docUserWorkbook = load_workbook(RESOURCES.excelList[numExcelMerged])
                docUserSheet = docUserWorkbook.active


                #Setting row, column parameters
                maxUserRow = docUserSheet.max_row + 1
                maxColumn = docUserSheet.max_column + 1
    
                docBuildSheet.column_dimensions["A"].width = 20
                docBuildSheet.column_dimensions["B"].width = 60
                docBuildSheet.column_dimensions["C"].width = 15
                docBuildSheet.column_dimensions["D"].width = 15
                docBuildSheet.column_dimensions["E"].width = 15
                docBuildSheet.column_dimensions["F"].width = 25
    
    
                #Reading and writing from an excel
                for i in range(1, maxUserRow):
                    userRow = str(i)
                    row = []        
                    for j in range(1, maxColumn):
                        column = chr(64+j)
                        userCell = column + userRow
                        row.append(docUserSheet[userCell].value)
                    row.append(GetExcelFileName(RESOURCES.excelList[numExcelMerged]))
                    
                    if(IfValue(row) == True):

                            if(type(row[0]) == str):
                                row[0] = DATETIME(row[0])
  
                            if(len(DATA) == 0):
                                DATA.append(row)
 
                            else:
                                for i in range (0, len(DATA)):
                                        
                                    if(row[0] <= DATA[i][0]):
                                        DATA.insert(i,row)
                                        break

                    else:
                        headingValue = headingValue + 1
                        
                        if(headingValue == 1):
                            row[-1] = "Source File"
                            docBuildSheet.append(row)                


                numExcelMerged = numExcelMerged + 1
                
            for i in range (0, len(DATA)):
                docBuildSheet.append(DATA[i])
                
            RESOURCES.excelList.clear()

            #Saving the Build File
            docBuildWorkbook.save(RESOURCES.buildFileNameExt)
        
            #Clearing Display
            RESOURCES.excelDisplay.clear()
            
            self.AfterMergeWindow.show()
            
    def closeEvent(self, event):
        self.error.close()
        self.AfterMergeWindow.close()